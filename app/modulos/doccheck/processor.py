import io
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill
from datetime import datetime

YELLOW_FILL = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")

class DocProcessor:
    """
    Procesa un Excel en memoria (BytesIO).
    Nunca escribe nada en disco.
    """

    def __init__(self, file_bytes: bytes):
        """
        Recibe los bytes del archivo Excel directamente.
        """
        self._raw = io.BytesIO(file_bytes)
        self._load()

    def _load(self):
        self._raw.seek(0)
        self.wb = load_workbook(self._raw)
        self.data_sheet = None

        for name in self.wb.sheetnames:
            if name.strip().upper() == "DATA":
                self.data_sheet = self.wb[name]
                break

        if not self.data_sheet:
            self.data_sheet = self.wb[self.wb.sheetnames[0]]

        self.rows = list(self.data_sheet.iter_rows(min_row=2))

        if "Historial" in self.wb.sheetnames:
            self.history = self.wb["Historial"]
        else:
            self.history = self.wb.create_sheet("Historial")
            self.history.append(["Registro", "Campo", "Antiguo", "Nuevo", "FechaHora"])

    def record_count(self):
        return len(self.rows)

    def get_record(self, idx):
        if idx < 0 or idx >= len(self.rows):
            return None

        row = self.rows[idx]
        val_fecha = row[11].value
        if isinstance(val_fecha, datetime):
            val_fecha = val_fecha.strftime("%d/%m/%Y")

        return {
            "N° Registro":    row[2].value,
            "Paquete":        row[1].value,
            "Caja":           row[0].value,
            "Tomo":           row[3].value,
            "Folios":         row[6].value,
            "Tipo Doc":       row[8].value,
            "N° Documento":   row[7].value,
            "Razón Social":   row[9].value,
            "RUC":            row[10].value,
            "Fecha Extrema":  val_fecha,
            "Observaciones":  row[12].value,
            "X1":             row[13].value,
            "X2":             row[14].value,
            "X3":             row[15].value,
        }

    def save_record(self, idx, new_values):
        if idx < 0 or idx >= len(self.rows):
            return False

        row = self.rows[idx]
        registro = row[2].value
        any_change = False

        mapping = {
            "Folios":         7,
            "N° Documento":   8,
            "Tipo Doc":       9,
            "Razón Social":  10,
            "RUC":           11,
            "Fecha Extrema": 12,
            "Observaciones": 13,
            "X1":            14,
            "X2":            15,
            "X3":            16,
        }

        for campo, col in mapping.items():
            tuple_idx = col - 1
            cell_obj = row[tuple_idx]
            old = cell_obj.value
            old = "" if old is None else old

            new = new_values.get(campo)
            new = "" if new is None else new

            if campo == "Fecha Extrema":
                if isinstance(old, datetime):
                    old_fmt = old.strftime("%d/%m/%Y")
                else:
                    old_fmt = str(old)
                new_fmt = str(new)

                if old_fmt != new_fmt:
                    try:
                        dt = datetime.strptime(new_fmt, "%d/%m/%Y")
                        write_val = dt
                    except ValueError:
                        write_val = new_fmt

                    any_change = True
                    cell = self.data_sheet.cell(row=idx + 2, column=col)
                    cell.value = write_val
                    cell.fill = YELLOW_FILL

                    ahora = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
                    self.history.append([registro, campo, old_fmt, new_fmt, ahora])
                continue

            if str(old) != str(new):
                write_val = new or None
                any_change = True

                cell = self.data_sheet.cell(row=idx + 2, column=col)
                cell.value = write_val
                cell.fill = YELLOW_FILL

                ahora = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
                self.history.append([registro, campo, old, new, ahora])

        if any_change:
            # Guardar los cambios de vuelta en el BytesIO interno
            self._raw = io.BytesIO()
            self.wb.save(self._raw)

        return any_change

    def get_workbook_bytes(self) -> bytes:
        """Devuelve los bytes actuales del workbook en memoria (para persistir en DB)."""
        buf = io.BytesIO()
        self.wb.save(buf)
        buf.seek(0)
        return buf.read()

    def get_last_change(self) -> dict | None:
        """Devuelve el último cambio registrado en el Historial, o None si no hay."""
        rows = list(self.history.iter_rows(min_row=2, values_only=True))
        if not rows:
            return None
        last = rows[-1]
        return {
            "registro": str(last[0]) if last[0] is not None else "",
            "campo":    str(last[1]) if last[1] is not None else "",
            "antiguo":  str(last[2]) if last[2] is not None else "",
            "nuevo":    str(last[3]) if last[3] is not None else "",
            "fecha":    str(last[4]) if last[4] is not None else "",
        }

    def export_excel_bytes(self, caja=None) -> bytes:
        """Genera el Excel exportado y devuelve los bytes directamente."""
        wb_new = Workbook()
        ws = wb_new.active
        ws.title = "Exportados"

        headers = [
            "N° DE CAJA", "N° DE PAQUETE", "N° DE REGISTRO", "TOMO",
            "RANGO INICIAL", "RANGO FINAL", "FOLIOS", "N ° DE DOCUMENTO",
            "TIPO DOCUMENTO", "RAZON SOCIAL", "RUC", "FECHA EXTREMA",
            "OBSERVACIONES", "X1(REC)", "X2(RC)", "X3"
        ]
        ws.append(headers)

        for row in self.rows:
            val_caja = row[0].value
            if caja and str(val_caja) != str(caja):
                continue

            val_fecha = row[11].value
            if isinstance(val_fecha, datetime):
                fecha = val_fecha.strftime("%d/%m/%Y")
            else:
                fecha = str(val_fecha)

            data = [
                row[0].value, row[1].value, row[2].value, row[3].value,
                "", "", row[6].value, row[7].value, row[8].value,
                row[9].value, row[10].value, fecha, row[12].value,
                row[13].value, row[14].value, row[15].value,
            ]
            ws.append(data)

        buf = io.BytesIO()
        wb_new.save(buf)
        buf.seek(0)
        return buf.read()

    def export_txt_bytes(self, caja=None) -> bytes:
        """Genera el TXT exportado y devuelve los bytes directamente."""
        headers = [
            "N° DE CAJA", "N° DE PAQUETE", "N° DE REGISTRO", "TOMO",
            "RANGO INICIAL", "RANGO FINAL", "FOLIOS", "N ° DE DOCUMENTO",
            "TIPO DOCUMENTO", "RAZON SOCIAL", "RUC", "FECHA EXTREMA",
            "OBSERVACIONES", "X1(REC)", "X2(RC)", "X3"
        ]

        lines = ["\t".join(headers)]

        for row in self.rows:
            val_caja = row[0].value
            if caja and str(val_caja) != str(caja):
                continue

            val_fecha = row[11].value
            if isinstance(val_fecha, datetime):
                fecha = val_fecha.strftime("%d/%m/%Y")
            else:
                fecha = str(val_fecha)
            if fecha == "None":
                fecha = ""

            values = [
                str(row[0].value or ""), str(row[1].value or ""),
                str(row[2].value or ""), str(row[3].value or ""),
                "", "", str(row[6].value or ""), str(row[7].value or ""),
                str(row[8].value or ""), str(row[9].value or ""),
                str(row[10].value or ""), fecha, str(row[12].value or ""),
                str(row[13].value or ""), str(row[14].value or ""),
                str(row[15].value or ""),
            ]
            lines.append("\t".join(values))

        return "\n".join(lines).encode("utf-8")
