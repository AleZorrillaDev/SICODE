"""
BPSearch Processor - Búsqueda de palabras en PDFs del Boletín El Peruano
"""
import os
import re
import fitz  # PyMuPDF
import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo
from datetime import datetime
import time
from typing import List, Dict, Optional, Generator
import threading


class BPSearchProcessor:
    """Procesador para búsqueda de palabras en PDFs"""
    
    def __init__(self):
        self.carpeta_principal = ""
        self.resultados: List[Dict] = []
        self.pausado = False
        self.cancelado = False
        self.progreso_actual = 0
        self.total_archivos = 0
        self.tiempo_inicio = 0
        self.archivo_actual = ""
        self.estado = "idle"  # idle, processing, paused, completed, cancelled
        
    def set_carpeta(self, carpeta: str) -> bool:
        """Establece la carpeta principal de búsqueda"""
        if os.path.isdir(carpeta):
            self.carpeta_principal = carpeta
            return True
        return False
    
    def _normalizar_texto(self, texto: str) -> str:
        """Quita tildes y normaliza texto"""
        reemplazos = {
            'á': 'a', 'Á': 'a',
            'é': 'e', 'É': 'e', 
            'í': 'i', 'Í': 'i',
            'ó': 'o', 'Ó': 'o',
            'ú': 'u', 'Ú': 'u',
            'ñ': 'n', 'Ñ': 'n'
        }
        for orig, reemplazo in reemplazos.items():
            texto = texto.replace(orig, reemplazo)
        return texto.lower()
    
    def _cargar_historial(self, palabra: str) -> set:
        """Carga el historial de PDFs procesados para una palabra"""
        nombre_historial = f"pdfs_procesados_{palabra.lower().strip()}.txt"
        ruta_historial = os.path.join(self.carpeta_principal, nombre_historial)
        if not os.path.exists(ruta_historial):
            return set()
        with open(ruta_historial, "r", encoding="utf-8") as f:
            return set(line.strip() for line in f.readlines())
    
    def _guardar_historial(self, pdfs: List[str], palabra: str):
        """Guarda el historial de PDFs procesados"""
        nombre_historial = f"pdfs_procesados_{palabra.lower().strip()}.txt"
        ruta_historial = os.path.join(self.carpeta_principal, nombre_historial)
        with open(ruta_historial, "a", encoding="utf-8") as f:
            for pdf in pdfs:
                f.write(f"{pdf}\n")
    
    def borrar_historial(self, palabra: str) -> bool:
        """Borra el historial de una palabra"""
        nombre_historial = f"pdfs_procesados_{palabra.lower().strip()}.txt"
        ruta_historial = os.path.join(self.carpeta_principal, nombre_historial)
        if os.path.exists(ruta_historial):
            os.remove(ruta_historial)
            return True
        return False
    
    def obtener_pdfs(self, incluir_subcarpetas: bool = False) -> List[tuple]:
        """Obtiene la lista de PDFs en la carpeta"""
        pdfs = []
        if incluir_subcarpetas:
            for raiz, dirs, archivos in os.walk(self.carpeta_principal):
                for archivo in archivos:
                    if archivo.lower().endswith('.pdf'):
                        pdfs.append((raiz, archivo))
        else:
            for archivo in os.listdir(self.carpeta_principal):
                if archivo.lower().endswith('.pdf'):
                    pdfs.append((self.carpeta_principal, archivo))
        return pdfs
    
    def buscar_palabra(self, palabra: str, incluir_subcarpetas: bool = False, 
                       ignorar_historial: bool = False) -> Generator[Dict, None, None]:
        """
        Busca una palabra en los PDFs de forma iterativa (generador).
        Permite pausar/cancelar y reporta progreso.
        """
        self.resultados = []
        self.pausado = False
        self.cancelado = False
        self.estado = "processing"
        self.tiempo_inicio = time.time()
        
        # Cargar historial si no se ignora
        historial = set() if ignorar_historial else self._cargar_historial(palabra)
        
        # Obtener PDFs
        todos_pdfs = self.obtener_pdfs(incluir_subcarpetas)
        pdfs_nuevos = [(c, a) for c, a in todos_pdfs if a not in historial]
        
        self.total_archivos = len(pdfs_nuevos)
        self.progreso_actual = 0
        
        if self.total_archivos == 0:
            self.estado = "completed"
            yield {
                "type": "complete",
                "message": "No se encontraron archivos PDF nuevos",
                "total": 0,
                "encontrados": 0
            }
            return
        
        # Preparar patrón de búsqueda
        palabra_normalizada = self._normalizar_texto(palabra)
        patron = re.compile(
            rf"(?i)(\b{re.escape(palabra_normalizada)}\b|"
            rf"{re.escape(palabra_normalizada[:-1])}-\s*\n\s*{re.escape(palabra_normalizada[-1])}\b)"
        )
        
        pdfs_procesados = []
        
        for i, (carpeta, nombre_pdf) in enumerate(pdfs_nuevos):
            # Verificar pausa
            while self.pausado and not self.cancelado:
                time.sleep(0.1)
            
            # Verificar cancelación
            if self.cancelado:
                self.estado = "cancelled"
                yield {
                    "type": "cancelled",
                    "message": "Búsqueda cancelada",
                    "procesados": i,
                    "total": self.total_archivos
                }
                break
            
            self.progreso_actual = i + 1
            self.archivo_actual = nombre_pdf
            ruta_pdf = os.path.join(carpeta, nombre_pdf)
            
            # Calcular tiempo estimado
            tiempo_transcurrido = time.time() - self.tiempo_inicio
            if i > 0:
                tiempo_estimado = int((tiempo_transcurrido / i) * (self.total_archivos - i))
            else:
                tiempo_estimado = 0
            
            # Reportar progreso
            yield {
                "type": "progress",
                "current": i + 1,
                "total": self.total_archivos,
                "percent": int((i + 1) / self.total_archivos * 100),
                "archivo": nombre_pdf,
                "tiempo_transcurrido": int(tiempo_transcurrido),
                "tiempo_estimado": tiempo_estimado
            }
            
            # Buscar en PDF
            try:
                doc = fitz.open(ruta_pdf)
                apariciones = 0
                paginas_encontradas = []
                
                for num_pagina, pagina in enumerate(doc):
                    texto = pagina.get_text("text") or ""
                    # Unir palabras cortadas por guion
                    texto = re.sub(r'(\w)-\s*\n\s*(\w)', r'\1\2', texto)
                    texto_normalizado = self._normalizar_texto(texto)
                    coincidencias = re.findall(patron, texto_normalizado)
                    cantidad = len(coincidencias)
                    if cantidad > 0:
                        apariciones += cantidad
                        paginas_encontradas.append(num_pagina + 1)
                
                doc.close()
                
                if apariciones > 0:
                    resultado = {
                        "ruta": ruta_pdf,
                        "nombre": nombre_pdf,
                        "veces": apariciones,
                        "paginas": paginas_encontradas
                    }
                    self.resultados.append(resultado)
                    
                    yield {
                        "type": "found",
                        "archivo": nombre_pdf,
                        "veces": apariciones,
                        "paginas": paginas_encontradas
                    }
                
                pdfs_procesados.append(nombre_pdf)
                
            except Exception as e:
                yield {
                    "type": "error",
                    "archivo": nombre_pdf,
                    "error": str(e)
                }
        
        # Guardar historial
        if pdfs_procesados and not self.cancelado:
            self._guardar_historial(pdfs_procesados, palabra)
        
        if not self.cancelado:
            self.estado = "completed"
            yield {
                "type": "complete",
                "message": "Búsqueda completada",
                "total": self.total_archivos,
                "encontrados": len(self.resultados),
                "tiempo_total": int(time.time() - self.tiempo_inicio)
            }
    
    def pausar(self):
        """Pausa la búsqueda"""
        self.pausado = True
        self.estado = "paused"
    
    def continuar(self):
        """Continúa la búsqueda"""
        self.pausado = False
        self.estado = "processing"
    
    def cancelar(self):
        """Cancela la búsqueda"""
        self.cancelado = True
        self.estado = "cancelled"
    
    def get_estado(self) -> Dict:
        """Retorna el estado actual del procesador"""
        return {
            "estado": self.estado,
            "progreso": self.progreso_actual,
            "total": self.total_archivos,
            "archivo_actual": self.archivo_actual,
            "resultados_count": len(self.resultados)
        }
    
    def exportar_txt(self, palabra: str) -> str:
        """Exporta resultados a TXT y retorna el contenido"""
        if not self.resultados:
            return ""
        
        lineas = [f"RESULTADOS DE LA BÚSQUEDA DE '{palabra.upper()}'\n"]
        lineas.append(f"Fecha: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        lineas.append(f"Total encontrados: {len(self.resultados)} archivos\n")
        lineas.append("=" * 60 + "\n\n")
        
        for r in self.resultados:
            lineas.append(f"Archivo: {r['nombre']}\n")
            lineas.append(f"  Ruta: {r['ruta']}\n")
            lineas.append(f"  Veces encontradas: {r['veces']}\n")
            lineas.append(f"  Páginas: {', '.join(map(str, r['paginas']))}\n\n")
        
        return "".join(lineas)
    
    def exportar_excel(self, palabra: str) -> bytes:
        """Exporta resultados a Excel y retorna los bytes"""
        from io import BytesIO
        
        if not self.resultados:
            return b""
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Resultados"
        
        encabezados = ["DIA", "MES", "AÑO", "BOLETIN", "VECES ENCONTRADO", "PÁGINAS"]
        ws.append(encabezados)
        
        for r in self.resultados:
            nombre_pdf = r['nombre']
            # Extraer fecha del nombre (formato: BOyyyymmdd.pdf)
            try:
                base = os.path.splitext(nombre_pdf)[0]
                if base.startswith("BO") and len(base) >= 10:
                    anio = base[2:6]
                    mes = base[6:8]
                    dia = base[8:10]
                else:
                    anio = mes = dia = ""
            except:
                anio = mes = dia = ""
            
            ws.append([
                dia,
                mes,
                anio,
                nombre_pdf,
                r["veces"],
                ", ".join(map(str, r["paginas"]))
            ])
        
        # Aplicar estilo de tabla
        num_rows = len(self.resultados) + 1
        tabla = Table(displayName="TablaResultados", ref=f"A1:F{num_rows}")
        estilo = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )
        tabla.tableStyleInfo = estilo
        ws.add_table(tabla)
        
        # Ajustar anchos de columna
        for columna in ws.columns:
            max_length = 0
            columna_letra = columna[0].column_letter
            for celda in columna:
                try:
                    if celda.value:
                        max_length = max(max_length, len(str(celda.value)))
                except:
                    pass
            ws.column_dimensions[columna_letra].width = max_length + 2
        
        # Guardar en memoria
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    def exportar_pdf(self, palabra: str) -> bytes:
        """Exporta resultados a PDF y retorna los bytes"""
        try:
            from fpdf import FPDF
        except ImportError:
            return b""
            
        if not self.resultados:
            return b""
            
        pdf = FPDF()
        pdf.add_page()
        pdf.set_font("Arial", size=12)
        pdf.multi_cell(0, 10, f"RESULTADOS DE LA BÚSQUEDA DE '{palabra.upper()}'\n\n")
        
        for r in self.resultados:
            pdf.set_font("Arial", "B", 11)
            # Manejo de encoding para fpdf (latin-1)
            try:
                ruta_str = r['ruta'].encode('latin-1', 'replace').decode('latin-1')
            except:
                ruta_str = r['ruta']
                
            pdf.multi_cell(0, 8, ruta_str)
            
            pdf.set_font("Arial", size=11)
            pdf.cell(0, 8, f"  Veces encontradas: {r['veces']}", ln=1)
            pdf.cell(0, 8, f"  Páginas: {', '.join(map(str, r['paginas']))}", ln=1)
            pdf.ln(2)
            
        try:
            return pdf.output(dest='S').encode('latin-1')
        except:
            return b""
