"""
Módulo Agregar - Sistema de Registro de Formularios SUNAT
Lógica de negocio: escritura en Excel usando openpyxl (+ xlwings en tiempo real si está disponible).
Incluye gestión de la ruta activa del Excel (con persistencia en JSON).
"""

import openpyxl
import os
import json

# ── Columnas del Excel ────────────────────────────────────────────────────────
HEADERS = [
    'N° CAJA', 'N° PAQUETE', 'NUMERO DE REGISTRO', 'TOMO',
    'RANGO INICIAL', 'RANGO FINAL', 'FOLIOS', 'TIPO DOCUMENTAL',
    'N° DE DOCUMENTO', 'RAZON SOCIAL', 'RUC', 'FECHA EXTREMA',
    'OBSERVACIONES', 'X1', 'X2', 'X3'
]

# xlwings es opcional (solo disponible en Windows con Excel abierto)
try:
    import xlwings as xw
    HAS_XLWINGS = True
except ImportError:
    HAS_XLWINGS = False

# ── Persistencia de la ruta activa ────────────────────────────────────────────
# Se guarda en un JSON junto al código del módulo para sobrevivir reinicios.
_CONFIG_FILE = os.path.join(os.path.dirname(__file__), "agregar_config.json")

def _default_excel_path() -> str:
    """Ruta por defecto si no hay configuración guardada."""
    # Directorio de trabajo (raíz del proyecto)
    root = os.path.abspath(".")
    return os.path.join(root, "FORMULARIO194_MESADEPARTES.xlsx")

def get_active_excel_path() -> str:
    """Lee la ruta del Excel activo desde la config persistida (o usa el default)."""
    try:
        if os.path.exists(_CONFIG_FILE):
            with open(_CONFIG_FILE, "r", encoding="utf-8") as f:
                cfg = json.load(f)
            path = cfg.get("excel_path", "")
            if path and os.path.exists(path):
                return path
    except Exception:
        pass
    return _default_excel_path()

def set_active_excel_path(path: str) -> None:
    """Guarda la ruta del Excel activo en la config persistida."""
    with open(_CONFIG_FILE, "w", encoding="utf-8") as f:
        json.dump({"excel_path": path}, f, ensure_ascii=False, indent=2)

# ── Helpers de Excel ──────────────────────────────────────────────────────────
def ensure_excel(excel_path: str):
    """Crea el archivo Excel con las hojas necesarias si no existe."""
    if os.path.exists(excel_path):
        return
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "FORMULARIO 194"
    for i, h in enumerate(HEADERS, 1):
        ws1.cell(1, i, h)
    ws2 = wb.create_sheet("EXPEDIENTES")
    for i, h in enumerate(HEADERS, 1):
        ws2.cell(1, i, h)
    wb.save(excel_path)


def _get_xlwings_book(excel_path: str):
    """Intenta encontrar el libro abierto en Excel via xlwings."""
    if not HAS_XLWINGS:
        return None
    try:
        fname = os.path.basename(excel_path)
        for a in xw.apps:
            for b in a.books:
                if b.name.lower() == fname.lower():
                    return b
    except Exception:
        pass
    return None


def write_record(excel_path: str, sheet_name: str, data: dict) -> dict:
    """
    Escribe un registro en la hoja indicada.
    Intenta xlwings (Excel en vivo) primero; si no está disponible, usa openpyxl.
    """
    live = False
    row = None
    book = _get_xlwings_book(excel_path)

    if book:
        try:
            ws = book.sheets[sheet_name]
            row = 2 if not ws.range("H2").value else ws.range("H1").end("down").row + 1
            for col, val in data.items():
                ws.range((row, col)).value = val
            live = True
        except Exception:
            live = False

    if not live:
        wb = openpyxl.load_workbook(excel_path)
        ws = wb[sheet_name]
        row = ws.max_row + 1
        for col, val in data.items():
            ws.cell(row, col, val)
        wb.save(excel_path)
        wb.close()

    return {"success": True, "live": live, "row": row}


def count_records(excel_path: str, sheet_name: str) -> int:
    """Cuenta los registros en la hoja indicada (descuenta la fila de encabezado)."""
    book = _get_xlwings_book(excel_path)
    if book:
        try:
            ws = book.sheets[sheet_name]
            if not ws.range("H2").value:
                return 0
            return max(0, ws.range("H1").end("down").row - 1)
        except Exception:
            pass
    try:
        wb = openpyxl.load_workbook(excel_path, read_only=True)
        ws = wb[sheet_name]
        n = max(0, ws.max_row - 1) if ws.max_row else 0
        wb.close()
        return n
    except Exception:
        return 0


def get_status(excel_path: str) -> dict:
    """Devuelve el estado de conexión con Excel, los conteos y la ruta activa."""
    book = _get_xlwings_book(excel_path)
    return {
        "excel_connected": book is not None,
        "has_xlwings": HAS_XLWINGS,
        "formulario_count": count_records(excel_path, "FORMULARIO 194"),
        "expediente_count": count_records(excel_path, "EXPEDIENTES"),
        "excel_path": excel_path,
        "excel_filename": os.path.basename(excel_path),
        "excel_exists": os.path.exists(excel_path),
    }
