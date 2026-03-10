from fastapi import APIRouter, Request, Depends, UploadFile, File, Form
from fastapi.responses import HTMLResponse, JSONResponse
from fastapi.templating import Jinja2Templates
from pydantic import BaseModel, field_validator
import os
import shutil
import tempfile

from app.core.config import settings
from app.modulos.auth.dependencies import login_required
from app.modulos.agregar.agregar_service import (
    ensure_excel, write_record, get_status, count_records,
    get_active_excel_path, set_active_excel_path
)

# ── Router ────────────────────────────────────────────────────────────────────
router = APIRouter(
    prefix="/agregar",
    tags=["Agregar"],
    dependencies=[Depends(login_required)]
)

# ── Templates ─────────────────────────────────────────────────────────────────
templates_dir    = settings.get_template_path("app/modulos/agregar/templates")
templates_base   = settings.get_template_path("app/modulos/inicio/templates")
templates_global = settings.get_template_path("app/templates")

templates = Jinja2Templates(directory=[templates_dir, templates_base, templates_global])


# ── Modelos Pydantic ──────────────────────────────────────────────────────────
class RegistroFormulario(BaseModel):
    documento: str
    razon_social: str = ""
    ruc: str
    fecha: str = ""
    observaciones: str = ""

    @field_validator("ruc")
    @classmethod
    def ruc_must_be_11(cls, v: str) -> str:
        v = v.strip()
        if len(v) != 11 or not v.isdigit():
            raise ValueError("RUC debe tener exactamente 11 dígitos")
        return v


class RegistroExpediente(BaseModel):
    parte1: str
    parte2: str
    parte3: str
    parte4: str
    razon_social: str = ""
    ruc: str
    fecha: str = ""
    observaciones: str = ""

    @field_validator("ruc")
    @classmethod
    def ruc_must_be_11(cls, v: str) -> str:
        v = v.strip()
        if len(v) != 11 or not v.isdigit():
            raise ValueError("RUC debe tener exactamente 11 dígitos")
        return v


# ── Página principal ──────────────────────────────────────────────────────────
@router.get("/", response_class=HTMLResponse)
async def agregar_home(request: Request):
    excel_path = get_active_excel_path()
    ensure_excel(excel_path)
    return templates.TemplateResponse("agregar/index.html", {"request": request})


# ── API: Status ───────────────────────────────────────────────────────────────
@router.get("/api/status")
async def api_status():
    return get_status(get_active_excel_path())


# ── API: Configurar Excel por ruta de servidor ────────────────────────────────
@router.post("/api/set-path")
async def set_excel_path(path: str = Form(...)):
    """Establece la ruta del archivo Excel existente en el servidor."""
    path = path.strip()
    if not path:
        return JSONResponse(status_code=400, content={"success": False, "error": "Ruta vacía."})
    if not os.path.exists(path):
        return JSONResponse(status_code=400, content={"success": False, "error": f"Archivo no encontrado: {path}"})
    if not path.lower().endswith((".xlsx", ".xls")):
        return JSONResponse(status_code=400, content={"success": False, "error": "El archivo debe ser .xlsx o .xls"})
    set_active_excel_path(path)
    return {
        "success": True,
        "excel_path": path,
        "excel_filename": os.path.basename(path),
    }


# ── API: Subir Excel desde el navegador ───────────────────────────────────────
@router.post("/api/upload-excel")
async def upload_excel(file: UploadFile = File(...)):
    """Recibe un archivo Excel subido desde el navegador y lo guarda como archivo activo."""
    if not file.filename.lower().endswith((".xlsx", ".xls")):
        return JSONResponse(status_code=400, content={"success": False, "error": "Solo se admiten archivos .xlsx o .xls"})

    # Guardar en la raíz del proyecto con el nombre original
    root = settings.get_app_root()
    dest_path = os.path.join(root, file.filename)

    content = await file.read()
    with open(dest_path, "wb") as f:
        f.write(content)

    set_active_excel_path(dest_path)
    return {
        "success": True,
        "excel_path": dest_path,
        "excel_filename": file.filename,
        "size_kb": round(len(content) / 1024, 1),
    }


# ── API: Crear nuevo Excel vacío ──────────────────────────────────────────────
@router.post("/api/new-excel")
async def new_excel(filename: str = Form("FORMULARIO194_MESADEPARTES.xlsx")):
    """Crea un nuevo Excel vacío con las hojas requeridas."""
    filename = filename.strip()
    if not filename.lower().endswith(".xlsx"):
        filename += ".xlsx"
    root = settings.get_app_root()
    dest_path = os.path.join(root, filename)
    ensure_excel(dest_path)          # solo crea si no existe
    # Si ya existía, lo sobreescribimos
    import openpyxl, json as _json
    from app.modulos.agregar.agregar_service import HEADERS
    wb = openpyxl.Workbook()
    ws1 = wb.active; ws1.title = "FORMULARIO 194"
    for i, h in enumerate(HEADERS, 1): ws1.cell(1, i, h)
    ws2 = wb.create_sheet("EXPEDIENTES")
    for i, h in enumerate(HEADERS, 1): ws2.cell(1, i, h)
    wb.save(dest_path)
    set_active_excel_path(dest_path)
    return {"success": True, "excel_path": dest_path, "excel_filename": filename}


# ── API: Guardar Formulario 194 ───────────────────────────────────────────────
@router.post("/api/formulario")
async def save_formulario(reg: RegistroFormulario):
    excel_path = get_active_excel_path()
    ensure_excel(excel_path)
    fecha_val = f"'{reg.fecha.strip()}" if reg.fecha.strip() else ""
    data = {
        8:  "FORMULARIO NRO. 194",
        9:  reg.documento.strip(),
        10: reg.razon_social.strip(),
        11: reg.ruc.strip(),
        12: fecha_val,
        13: reg.observaciones.strip(),
    }
    result = write_record(excel_path, "FORMULARIO 194", data)
    result["count"] = count_records(excel_path, "FORMULARIO 194")
    return result


# ── API: Guardar Expediente ───────────────────────────────────────────────────
@router.post("/api/expediente")
async def save_expediente(reg: RegistroExpediente):
    excel_path = get_active_excel_path()
    ensure_excel(excel_path)
    fecha_val = f"'{reg.fecha.strip()}" if reg.fecha.strip() else ""
    doc = f"000-URD{reg.parte1.strip()}-{reg.parte2.strip()}-{reg.parte3.strip()}-{reg.parte4.strip()}"
    data = {
        8:  "EXPEDIENTE DE MEZA DE PARTES",
        9:  doc,
        10: reg.razon_social.strip(),
        11: reg.ruc.strip(),
        12: fecha_val,
        13: reg.observaciones.strip(),
    }
    result = write_record(excel_path, "EXPEDIENTES", data)
    result["count"] = count_records(excel_path, "EXPEDIENTES")
    return result
